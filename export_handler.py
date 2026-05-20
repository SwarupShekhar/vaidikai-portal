import json
import os
import io
import time
import shutil
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional

import cv2
import numpy as np
import pandas as pd
import requests
from azure.storage.blob import BlobServiceClient
from dotenv import load_dotenv
from collateral_detector import generate_signature, find_duplicates, store_signatures
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

load_dotenv()

def _get_ls_headers() -> dict:
    """Helper to get authenticated headers for Label Studio API."""
    ls_url = os.getenv("LABEL_STUDIO_URL", "").rstrip("/")
    api_key = os.getenv("LABEL_STUDIO_API_KEY", "")
    
    if not ls_url or not api_key:
        raise ValueError("LABEL_STUDIO_URL or LABEL_STUDIO_API_KEY not found in environment")

    # If it's a refresh token (JWT starting with eyJ), exchange it
    if api_key.startswith("eyJ"):
        import base64
        try:
            payload = api_key.split(".")[1]
            payload += "=" * (4 - len(payload) % 4)
            claims = json.loads(base64.b64decode(payload))
            if claims.get("token_type") == "refresh":
                r = requests.post(f"{ls_url}/api/token/refresh", json={"refresh": api_key}, timeout=15)
                r.raise_for_status()
                access_token = r.json()["access"]
                return {"Authorization": f"Bearer {access_token}"}
        except Exception as e:
            print(f"Token refresh failed, using as-is: {e}")
            
    return {"Authorization": f"Bearer {api_key}"}

def export_and_deliver(
    client_code: str,
    original_filename: str,
    project_id: str,
    label_id: str = None,
    task_id: int = None,
    force_delivery: bool = False,
    internal_export: bool = False,
) -> Dict[str, Any]:
    """
    Export completed annotations from Label Studio and deliver to client via Azure Blob.
    For jewelry projects, generates a Gold Standard Bundle (.zip) with audit images,
    COCO JSON, and runs duplicate collateral detection before delivery.
    """
    try:
        print(f"Starting XLSX export and delivery for {client_code}/{original_filename}")

        ls_url = os.getenv("LABEL_STUDIO_URL", "").rstrip("/")
        connection_string = os.getenv("AZURE_STORAGE_CONNECTION_STRING")

        if not connection_string:
            raise ValueError("AZURE_STORAGE_CONNECTION_STRING not found in environment")

        headers = _get_ls_headers()

        # STEP 1: Fetch Task(s)
        matched_tasks = []
        if task_id:
            print(f"Fetching specific task ID: {task_id}")
            r = requests.get(f"{ls_url}/api/tasks/{task_id}", headers=headers, timeout=15)
            if r.status_code == 200:
                matched_tasks = [r.json()]
        
        if not matched_tasks:
            import time
            global _TASK_CACHE
            project_ids_to_check = [str(project_id)]
            for pid in ["1", "2", "3", "4", "5", "6", "7"]:
                if pid not in project_ids_to_check:
                    project_ids_to_check.append(pid)

            for pid in project_ids_to_check:
                try:
                    now = time.time()
                    export_cache_key = f"export_{ls_url}_{pid}"
                    tasks_cache_key = f"{ls_url}_{pid}"

                    if export_cache_key in _TASK_CACHE and now - _TASK_CACHE[export_cache_key]["time"] < 15:
                        all_tasks = _TASK_CACHE[export_cache_key]["tasks"]
                    else:
                        print(f"Fetching bulk JSON export for project {pid}...")
                        r = requests.get(
                            f"{ls_url}/api/projects/{pid}/export?exportType=JSON",
                            headers=headers,
                            timeout=20
                        )
                        all_tasks = r.json() if r.status_code == 200 else []
                        _TASK_CACHE[export_cache_key] = {"tasks": all_tasks, "time": now}

                    matched = [
                        t for t in all_tasks
                        if t.get("data", {}).get("client_code") == client_code
                        and t.get("data", {}).get("filename") == original_filename
                        and len(t.get("annotations", [])) > 0
                    ]
                    if not matched:
                        # Fallback to /api/tasks to catch unsubmitted/imported pre-annotations
                        if tasks_cache_key in _TASK_CACHE and now - _TASK_CACHE[tasks_cache_key]["time"] < 15:
                            tasks_list = _TASK_CACHE[tasks_cache_key]["tasks"]
                        else:
                            tr = requests.get(
                                f"{ls_url}/api/tasks",
                                params={"project": pid, "page_size": 1000, "expand": "annotations"},
                                headers=headers,
                                timeout=20
                            )
                            tasks_list = tr.json() if tr.status_code == 200 else []
                            if isinstance(tasks_list, dict):
                                tasks_list = tasks_list.get("tasks", [])
                            _TASK_CACHE[tasks_cache_key] = {"tasks": tasks_list, "time": now}

                        matched = [
                            t for t in tasks_list
                            if t.get("data", {}).get("client_code") == client_code
                            and t.get("data", {}).get("filename") == original_filename
                            and (t.get("total_annotations", 0) > 0 or len(t.get("annotations", [])) > 0)
                        ]
                        
                    if matched:
                        print(f"Found {len(matched)} candidate tasks in project {pid}")
                        matched_tasks = matched
                        project_id = pid
                        break
                except Exception as ex:
                    print(f"Error checking project {pid}: {ex}")

        if not matched_tasks:
            print(f"No annotated tasks found for {client_code}/{original_filename}")
            return {
                "status": "error",
                "message": f"No completed annotations found for {original_filename}",
                "client_code": client_code
            }

        print(f"Found {len(matched_tasks)} annotated task(s)")

        segments: List[Dict[str, Any]] = []
        annotators = set()
        languages = set()

        is_audio = str(project_id) == os.getenv("LABEL_STUDIO_PROJECT_ID", "1")
        is_jewelry = str(project_id) == os.getenv("LABEL_STUDIO_JEWELRY_PROJECT_ID", "2")
        is_form = str(project_id) == os.getenv("LABEL_STUDIO_FORM_PROJECT_ID", "3")
        is_clickstream = str(project_id) == os.getenv("LABEL_STUDIO_CLICKSTREAM_PROJECT_ID", "4")
        is_housing = str(project_id) == os.getenv("LABEL_STUDIO_HOUSING_PROJECT_ID", "5")
        is_business = str(project_id) == os.getenv("LABEL_STUDIO_BUSINESS_PROJECT_ID", "6")
        is_transcript = str(project_id) == os.getenv("LABEL_STUDIO_TRANSCRIPT_PROJECT_ID", "7")
        is_image_project = is_jewelry or is_housing or is_business

        for task in matched_tasks:
            task_data = task.get("data", {})
            task_lang = task_data.get("language", "Unknown")
            languages.add(task_lang)
            
            for annotation in task.get("annotations", []):
                if annotation.get("was_cancelled"):
                    continue
                
                # Only export annotations explicitly accepted by a reviewer.
                # If no reviewer workflow is configured in Label Studio, all annotations
                # will be skipped here. Set up a review step in LS or remove this gate.
                if not annotation.get("is_accepted") and not is_clickstream:
                    print(f"[export] Skipping annotation id={annotation.get('id')} for {original_filename} — not accepted (is_accepted=False). Configure a reviewer workflow in Label Studio.")
                    continue
                
                result = annotation.get("result", [])

                if is_audio:
                    manual_status = "Pending"
                    for r_item in result:
                        if r_item.get("from_name") == "review_status":
                            manual_status = r_item.get("value", {}).get("choices", ["Pending"])[0]
                            break
                    review_result = annotation.get("review_result")
                    task_reviews = task.get("reviews", [])
                    

                email = annotation.get("created_username") or "Annotator"
                annotators.add(email.split("@")[0] if "@" in email else email)
                
                if is_image_project:
                    house_type = "N/A"
                    storey_count = "N/A"
                    nature_of_business = "N/A"
                    for r_item in result:
                        if r_item.get("type") == "choices":
                            fn = r_item.get("from_name", "")
                            c_vals = r_item.get("value", {}).get("choices", [])
                            if fn == "construction_type" and c_vals:
                                house_type = c_vals[0]
                            elif fn == "storey_count" and c_vals:
                                storey_count = c_vals[0]
                            elif fn == "business_type" and c_vals:
                                nature_of_business = c_vals[0]
                                
                    for r_item in result:
                        val = r_item.get("value", {})
                        rtype = r_item.get("type", "")
                        if rtype in ("polygonlabels", "keypointlabels", "rectanglelabels"):
                            default_cat = "General Jewelry"
                            if is_housing:
                                default_cat = "General House"
                            elif is_business:
                                default_cat = "General Business"
                                
                            labels = val.get(rtype, [default_cat])
                            category = labels[0] if labels else default_cat
                            points = val.get("points", [])
                            points_count = len(points) if rtype == "polygonlabels" else 1
                            segments.append({
                                "Category": category,
                                "Geometry Type": rtype.replace("labels", "").capitalize(),
                                "Points Count": points_count,
                                "Annotator": email.split("@")[0] if "@" in email else email,
                                "Image File": original_filename,
                                "Points": points,
                                "RType": rtype,
                                "House Type": house_type,
                                "Storey Count": storey_count,
                                "Nature of Business": nature_of_business
                            })
                elif is_clickstream:
                    session_status = "Smooth Journey"
                    friction_types = "None"
                    summary_text = ""
                    story_gist_text = ""
                    journey_intent = "Unknown"
                    journey_outcome = "Unknown"
                    root_cause = "None"
                    # Ami's expanded schema (2026-05-19)
                    event_class = "Page View"
                    granular_intent = "Product Exploration — Passive"
                    journey_stage = "Engagement"
                    product = "Multi-Product / General"
                    archetype = "Curious Browser"
                    # User-graph layer (Option D 2026-05-20)
                    user_archetype = "Single-Visit Drop"

                    for r_item in result:
                        fn = r_item.get("from_name", "")
                        val = r_item.get("value", {})
                        if fn == "session_status":
                            session_status = val.get("choices", [session_status])[0]
                        elif fn == "friction_types":
                            friction_types = ", ".join(val.get("choices", ["None"]))
                        elif fn == "summary":
                            summary_text = val.get("text", [""])[0]
                        elif fn == "story_gist":
                            story_gist_text = val.get("text", [""])[0]
                        elif fn == "journey_intent":
                            journey_intent = val.get("choices", [journey_intent])[0]
                        elif fn == "journey_outcome":
                            journey_outcome = val.get("choices", [journey_outcome])[0]
                        elif fn == "root_cause":
                            root_cause = val.get("choices", [root_cause])[0]
                        elif fn == "event_class":
                            event_class = val.get("choices", [event_class])[0]
                        elif fn == "granular_intent":
                            granular_intent = val.get("choices", [granular_intent])[0]
                        elif fn == "journey_stage":
                            journey_stage = val.get("choices", [journey_stage])[0]
                        elif fn == "product":
                            product = val.get("choices", [product])[0]
                        elif fn == "archetype":
                            archetype = val.get("choices", [archetype])[0]
                        elif fn == "user_archetype":
                            user_archetype = val.get("choices", [user_archetype])[0]

                    # Segmentation metadata carried in task.data (survives even
                    # if annotators don't touch the new controls).
                    segments.append({
                        "Session File": original_filename,
                        "Session ID": task_data.get("session_id", ""),
                        "User Type": task_data.get("user_type", "Unknown"),
                        "Device": task_data.get("device", "Unknown"),
                        "Device Cohort": task_data.get("device_cohort", ""),
                        "App Version": task_data.get("app_version", ""),
                        "Platform": task_data.get("platform", "Mobile"),
                        "Events": task_data.get("event_count", 0),
                        "Status": session_status,
                        "Friction Detected": friction_types,
                        "AI Summary": summary_text,
                        "Story Gist": story_gist_text or task_data.get("story_gist", ""),
                        "User Intent": journey_intent,
                        "Outcome": journey_outcome,
                        "Root Cause": root_cause,
                        # New label dimensions (Ami)
                        "Event Class": event_class,
                        "Granular Intent": granular_intent,
                        "Journey Stage": journey_stage,
                        "Product": product,
                        "Archetype": archetype,
                        # User-graph layer (Option D)
                        "User ID": task_data.get("user_id", ""),
                        "User Archetype": user_archetype,
                        "User Session #": task_data.get("user_session_index", 1),
                        "User Total Sessions": task_data.get("user_session_count", 1),
                        "User Recovery": "Yes" if task_data.get("user_recovery_flag") else "No",
                        "User Persistent Friction": ", ".join(task_data.get("user_persistent_friction") or []) or "—",
                        "Session Date": task_data.get("session_date", ""),
                    })
                elif is_form:
                    extracted_text = ""
                    for r_item in result:
                        fn = r_item.get("from_name", "")
                        val = r_item.get("value", {})
                        if fn == "extracted_text":
                            extracted_text = val.get("text", [""])[0]

                    segments.append({
                        "Form File": original_filename,
                        "Extracted OCR Text": extracted_text
                    })
                elif is_audio:
                    regions = {}
                    full_override = ""
                    for r_item in result:
                        fn = r_item.get("from_name", "")
                        val = r_item.get("value", {})
                        rtype = r_item.get("type")

                        # Free-form full-transcript override (authoritative when filled)
                        if fn == "full_transcript":
                            txts = val.get("text", [])
                            full_override = (txts[0] if txts else "").strip()
                            continue
                        # Non-region controls — not part of segment building
                        if fn == "review_status" or rtype == "choices":
                            continue

                        rid = r_item.get("id")
                        pid = r_item.get("parentID")
                        target_id = pid if pid else rid
                        if not target_id:
                            continue
                        if target_id.endswith("_t"):
                            target_id = target_id[:-2]

                        if target_id not in regions:
                            regions[target_id] = {
                                "start": val.get("start", 0),
                                "end": val.get("end", 0),
                                "speaker": "Unknown",
                                "transcript": "",
                                "language": task_lang
                            }

                        if rtype == "labels":
                            labels = val.get("labels", [])
                            regions[target_id]["speaker"] = labels[0] if labels else "Unknown"
                            regions[target_id]["start"] = val.get("start", regions[target_id]["start"])
                            regions[target_id]["end"] = val.get("end", regions[target_id]["end"])
                        elif rtype == "textarea":
                            texts = val.get("text", [])
                            regions[target_id]["transcript"] = texts[0] if texts else ""

                    if full_override:
                        # Annotator rewrote the whole transcript — it wins over
                        # the per-segment regions. Lines prefixed with
                        # [MM:SS - MM:SS] keep their timestamps; a line without
                        # the prefix inherits the previous line's end time.
                        import re as _re
                        _ts_re = _re.compile(
                            r'^\[\s*(\d+):(\d{2})\s*-\s*(\d+):(\d{2})\s*\]\s*([^:]{1,40}):\s*(.+)$')
                        _plain_re = _re.compile(r'^([^:]{1,40}):\s*(.+)$')
                        _prev_end = 0.0
                        for line in full_override.splitlines():
                            line = line.strip()
                            if not line:
                                continue
                            mt = _ts_re.match(line)
                            if mt:
                                st = int(mt.group(1)) * 60 + int(mt.group(2))
                                en = int(mt.group(3)) * 60 + int(mt.group(4))
                                spk = mt.group(5).strip()
                                txt = mt.group(6).strip()
                            else:
                                mp = _plain_re.match(line)
                                spk, txt = (mp.group(1).strip(), mp.group(2).strip()) if mp else ("Unknown", line)
                                st = _prev_end
                                en = _prev_end
                            dur = round(en - st, 3)
                            segments.append({
                                "Speaker": spk,
                                "Start Time (s)": round(st, 3),
                                "End Time (s)": round(en, 3),
                                "Duration (s)": dur if dur > 0 else 0,
                                "Transcript": txt,
                                "Language": task_lang,
                                "Audio File": original_filename
                            })
                            _prev_end = en
                    else:
                        for rid, data in regions.items():
                            if not data["transcript"].strip():
                                continue
                            duration = round(data["end"] - data["start"], 3)
                            segments.append({
                                "Speaker": data["speaker"],
                                "Start Time (s)": round(data["start"], 3),
                                "End Time (s)": round(data["end"], 3),
                                "Duration (s)": duration,
                                "Transcript": data["transcript"],
                                "Language": data["language"],
                                "Audio File": original_filename
                            })
                elif is_transcript:
                    dialogue = task_data.get("dialogue", [])
                    label_map = {}
                    for r_item in result:
                        if r_item.get("type") == "paragraphlabels":
                            val = r_item.get("value", {})
                            try:
                                start_idx = int(val.get("start", -1))
                                labels = val.get("paragraphlabels", [])
                                if start_idx != -1 and labels:
                                    label_map[start_idx] = labels[0]
                            except Exception:
                                pass
                    
                    for i, p_item in enumerate(dialogue):
                        speaker = label_map.get(i, p_item.get("author", "Unknown"))
                        text = p_item.get("text", "")
                        segments.append({
                            "Speaker": speaker,
                            "Transcript": text,
                            "Language": task_lang,
                            "Source File": original_filename
                        })

        final_segments = []
        call_insights = None  # populated only for is_audio (GPT-4o Call Insights sheet)
        clickstream_segments_breakdown = None  # populated only for is_clickstream
        clickstream_by_user_rows = None  # populated only for is_clickstream (Option D)
        if is_image_project:
            for i, seg in enumerate(segments):
                seg["Item #"] = i + 1
                row_data = {
                    "Item #": seg["Item #"],
                    "Category": seg["Category"],
                    "Geometry Type": seg["Geometry Type"],
                    "Points Count": seg["Points Count"],
                    "Image File": seg["Image File"]
                }
                if is_housing:
                    row_data["House Type"] = seg.get("House Type", "N/A")
                    row_data["Storey Count"] = seg.get("Storey Count", "N/A")
                if is_business:
                    row_data["Nature of Business"] = seg.get("Nature of Business", "N/A")
                if internal_export:
                    row_data["Annotator"] = seg["Annotator"]
                final_segments.append(row_data)

            if internal_export:
                columns = ["Item #", "Category", "Geometry Type", "Points Count", "Annotator", "Image File"]
            else:
                columns = ["Item #", "Category", "Geometry Type", "Points Count", "Image File"]
                
            if is_housing:
                columns.append("House Type")
                columns.append("Storey Count")
            if is_business:
                columns.append("Nature of Business")
            
            # Summary Configuration
            category_counts = {}
            for s in segments:
                category_counts[s["Category"]] = category_counts.get(s["Category"], 0) + 1
                
            overview_title = "Inventory Overview"
            if is_housing:
                overview_title = "Housing Validation Overview"
            elif is_business:
                overview_title = "MSME Business Overview"
                
            summary_headers = [overview_title, "Total Count"]
            summary_values_list = []
            for cat, count in sorted(category_counts.items()):
                summary_values_list.append([cat, count])
            summary_values_list.append(["Total Items", len(segments)])
            
            annotators_str = ", ".join(sorted(list(annotators))) or "Vaidik AI" if internal_export else "Vaidik AI"
            summary_values_list.append(["Annotated By", annotators_str])
            summary_values_list.append(["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
            
        elif is_clickstream:
            for i, seg in enumerate(segments):
                final_segments.append({
                    "Session #": i + 1,
                    "Session ID": seg.get("Session ID", ""),
                    "Session Date": seg.get("Session Date", ""),
                    # User-graph identity (Option D) — placed early so leadership
                    # readers see "who" before "what"
                    "User ID (hashed)": str(seg.get("User ID", ""))[:24] + ("…" if len(str(seg.get("User ID", ""))) > 24 else ""),
                    "User Session N/M": f"{seg.get('User Session #', 1)} / {seg.get('User Total Sessions', 1)}",
                    "User Archetype": seg.get("User Archetype", "Single-Visit Drop"),
                    "User Recovery": seg.get("User Recovery", "No"),
                    "Persistent Friction (across user)": seg.get("User Persistent Friction", "—"),
                    # Session-level dimensions
                    "User Type": seg.get("User Type", "Unknown"),
                    "Device": seg.get("Device", "Unknown"),
                    "App Version": seg.get("App Version", ""),
                    "Platform": seg.get("Platform", "Mobile"),
                    "Events": seg.get("Events", 0),
                    "Event Class": seg.get("Event Class", "Page View"),
                    "Journey Stage": seg.get("Journey Stage", "Engagement"),
                    "Product": seg.get("Product", "Multi-Product / General"),
                    "User Intent": seg["User Intent"],
                    "Granular Intent": seg.get("Granular Intent", ""),
                    "Outcome": seg["Outcome"],
                    "Status": seg["Status"],
                    "Friction Detected": seg["Friction Detected"],
                    "Root Cause": seg["Root Cause"],
                    "Archetype (session)": seg.get("Archetype", "Curious Browser"),
                    "AI Summary": seg["AI Summary"],
                    "Story Gist": seg.get("Story Gist", ""),
                    "Session File": seg["Session File"],
                })
            columns = [
                "Session #", "Session ID", "Session Date",
                "User ID (hashed)", "User Session N/M", "User Archetype", "User Recovery", "Persistent Friction (across user)",
                "User Type", "Device", "App Version", "Platform", "Events",
                "Event Class", "Journey Stage", "Product", "User Intent", "Granular Intent",
                "Outcome", "Status", "Friction Detected", "Root Cause", "Archetype (session)",
                "AI Summary", "Story Gist", "Session File",
            ]
            # === Executive KPIs (Summary sheet) ===
            from collections import Counter as _Counter
            n_total = len(segments) or 1  # avoid div-by-zero
            outcome_counts  = _Counter(s["Outcome"] for s in segments)
            stage_counts    = _Counter(s.get("Journey Stage", "Engagement") for s in segments)
            product_counts  = _Counter(s.get("Product", "Multi-Product / General") for s in segments)
            archetype_counts= _Counter(s.get("Archetype", "Curious Browser") for s in segments)
            usertype_counts = _Counter(s.get("User Type", "Unknown") for s in segments)
            cohort_counts   = _Counter(s.get("Device Cohort", "") for s in segments)
            root_counts     = _Counter(s["Root Cause"] for s in segments)
            friction_flat   = _Counter()
            for s in segments:
                for f in str(s.get("Friction Detected", "")).split(","):
                    f = f.strip()
                    if f and f.lower() not in ("", "none"):
                        friction_flat[f] += 1

            def _pct(n):
                return f"{(n/n_total*100):.1f}%"

            # Drop-off funnel by Journey Stage
            stage_order = ["Onboarding", "Engagement", "Transaction"]
            funnel_rows = []
            cumulative = n_total
            for st in stage_order:
                reached = sum(1 for s in segments
                              if stage_order.index(s.get("Journey Stage", "Engagement")) >= stage_order.index(st))
                funnel_rows.append([st, reached, _pct(reached)])

            # Build summary headers + value rows (sectioned)
            summary_headers = ["Metric", "Value", "% of Total"]
            summary_values_list = []
            summary_values_list.append(["— OVERVIEW —", "", ""])
            summary_values_list.append(["Total Sessions Analyzed", n_total, "100.0%"])
            summary_values_list.append(["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S"), ""])
            summary_values_list.append(["", "", ""])
            summary_values_list.append(["— DROP-OFF FUNNEL (by Journey Stage) —", "", ""])
            for row in funnel_rows:
                summary_values_list.append(row)
            summary_values_list.append(["", "", ""])
            summary_values_list.append(["— JOURNEY OUTCOME —", "", ""])
            for k, v in outcome_counts.most_common():
                summary_values_list.append([k, v, _pct(v)])
            summary_values_list.append(["", "", ""])
            summary_values_list.append(["— TOP FRICTION SIGNALS —", "", ""])
            for k, v in friction_flat.most_common(8):
                summary_values_list.append([k, v, _pct(v)])
            if not friction_flat:
                summary_values_list.append(["No friction detected in any session", "", ""])
            summary_values_list.append(["", "", ""])
            summary_values_list.append(["— ROOT CAUSE BREAKDOWN —", "", ""])
            for k, v in root_counts.most_common():
                summary_values_list.append([k, v, _pct(v)])
            summary_values_list.append(["", "", ""])
            summary_values_list.append(["— PRODUCT INTEREST —", "", ""])
            for k, v in product_counts.most_common():
                summary_values_list.append([k, v, _pct(v)])

            # === USER-LEVEL aggregations (Option D) ===
            # Collapse sessions to one row per unique hashed phone.
            users_map = {}  # uid -> {sessions: [seg], archetype, recovery, persistent}
            for s in segments:
                uid = s.get("User ID") or s.get("Session ID")
                if uid not in users_map:
                    users_map[uid] = {"sessions": [], "user_id": uid}
                users_map[uid]["sessions"].append(s)
                # User-level fields are constant across a user's sessions —
                # take from any session (latest wins on ties).
                users_map[uid]["user_archetype"] = s.get("User Archetype", "Single-Visit Drop")
                users_map[uid]["recovery"] = s.get("User Recovery", "No")
                users_map[uid]["persistent_friction"] = s.get("User Persistent Friction", "—")
                users_map[uid]["user_type"] = s.get("User Type", "Unknown")
                users_map[uid]["device"] = s.get("Device", "Unknown")

            n_users = len(users_map) or 1
            user_archetype_counts = _Counter(u["user_archetype"] for u in users_map.values())
            recovery_users = sum(1 for u in users_map.values() if u["recovery"] == "Yes")
            persistent_friction_users = sum(1 for u in users_map.values()
                                             if u["persistent_friction"] not in ("", "—", None))
            multi_session_users = sum(1 for u in users_map.values() if len(u["sessions"]) > 1)

            # User-level funnel — % of UNIQUE users who ever reached each stage
            user_max_stage = {}
            for uid, u in users_map.items():
                cur = -1
                for s in u["sessions"]:
                    try:
                        idx = stage_order.index(s.get("Journey Stage", "Engagement"))
                        if idx > cur:
                            cur = idx
                    except ValueError:
                        pass
                user_max_stage[uid] = cur
            user_funnel_rows = []
            for i, st in enumerate(stage_order):
                reached = sum(1 for v in user_max_stage.values() if v >= i)
                user_funnel_rows.append([st, reached, f"{reached/n_users*100:.1f}%"])

            # Inject user-level KPIs into the Summary sheet at the top
            user_kpi_block = [
                ["", "", ""],
                ["— USER-LEVEL VIEW (Option D: hashed-phone aggregation) —", "", ""],
                ["Unique Users Analysed", n_users, f"{n_users}/{n_total} sessions"],
                ["Users with Multiple Sessions", multi_session_users, f"{multi_session_users/n_users*100:.1f}%"],
                ["Recovery Rate (failed→succeeded)", recovery_users, f"{recovery_users/n_users*100:.1f}%"],
                ["Users with Persistent Friction (2+ sessions)", persistent_friction_users, f"{persistent_friction_users/n_users*100:.1f}%"],
                ["", "", ""],
                ["— USER-LEVEL FUNNEL (% of unique users who ever reached) —", "", ""],
            ]
            user_kpi_block.extend(user_funnel_rows)
            user_kpi_block.append(["", "", ""])
            user_kpi_block.append(["— USER ARCHETYPE DISTRIBUTION —", "", ""])
            for k, v in user_archetype_counts.most_common():
                user_kpi_block.append([k, v, f"{v/n_users*100:.1f}%"])
            # Append KPI block to summary_values_list
            summary_values_list.extend(user_kpi_block)

            # Hand off the segmentation breakdown to a dedicated sheet
            clickstream_segments_breakdown = {
                "User Type (CleverTap EP_USER_TYPE)": usertype_counts.most_common(),
                "Device · App-Version Cohort":        cohort_counts.most_common(10),
                "Behavioural Archetype (session)":    archetype_counts.most_common(),
                "User Archetype (cross-session)":     user_archetype_counts.most_common(),
            }

            # Build the "By User" sheet payload — one row per unique hashed phone.
            clickstream_by_user_rows = []
            outcome_short = {
                "Successfully Completed":   "✓ Done",
                "Abandoned Mid-Journey":    "✗ Abandoned",
                "Blocked by Error":         "✗ Error",
                "Deflected to Support":     "→ Support",
                "Browsing / Inconclusive":  "… Browsing",
            }
            # Sort users by total sessions desc — most-active users first
            sorted_users = sorted(users_map.items(),
                                  key=lambda kv: -len(kv[1]["sessions"]))
            for uid, u in sorted_users:
                user_sessions_ordered = sorted(u["sessions"],
                                                key=lambda s: int(s.get("User Session #", 1)))
                arc = " → ".join(
                    outcome_short.get(s["Outcome"], s["Outcome"][:10])
                    for s in user_sessions_ordered
                )
                clickstream_by_user_rows.append({
                    "User ID (hashed)": str(uid)[:32] + ("…" if len(str(uid)) > 32 else ""),
                    "Total Sessions": len(u["sessions"]),
                    "User Archetype": u["user_archetype"],
                    "Recovery": u["recovery"],
                    "Persistent Friction": u["persistent_friction"],
                    "User Type": u["user_type"],
                    "Device": u["device"],
                    "Session Arc (chronological)": arc,
                })
        elif is_form:
            for i, seg in enumerate(segments):
                final_segments.append({
                    "Scan #": i + 1,
                    "Form File": seg["Form File"],
                    "Extracted OCR Text": seg["Extracted OCR Text"]
                })
            columns = ["Scan #", "Form File", "Extracted OCR Text"]
            summary_headers = ["Form Processing Overview", "Total Count"]
            summary_values_list = [
                ["Total Forms Processed", len(segments)],
                ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
        elif is_transcript:
            for i, seg in enumerate(segments):
                row = {
                    "Turn #": i + 1,
                    "Speaker": seg["Speaker"],
                    "Transcript": seg["Transcript"],
                    "Language": seg["Language"],
                    "Source File": seg["Source File"]
                }
                final_segments.append(row)
            columns = ["Turn #", "Speaker", "Transcript", "Language", "Source File"]
            summary_headers = ["Transcript Overview", "Total Count"]
            summary_values_list = [
                ["Total Dialogue Turns", len(segments)],
                ["Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
        elif is_audio:
            segments.sort(key=lambda x: x.get("Start Time (s)", 0))
            for i, seg in enumerate(segments):
                seg["Segment #"] = i + 1

            lang_str = sorted(list(languages))[0] if languages else "Annotated"
            transcript_col = f"Transcript ({lang_str})"

            # Per-segment Intent/Sentiment/Outcome from the HUMAN-CORRECTED
            # text (rule-based engine reused from processor — cheap,
            # deterministic). Recomputed here, not from raw ASR pre-tags.
            try:
                from processor import (
                    tag as _tag, INTENT_RULES as _IR,
                    SENTIMENT_RULES as _SR, OUTCOME_RULES as _OR,
                )
            except Exception as _ie:
                print(f"[export] tag engine import failed: {_ie}")
                _tag = None

            for s in segments:
                txt = s["Transcript"]
                row = {
                    "Segment #": s["Segment #"],
                    "Speaker": s["Speaker"],
                    "Start Time (s)": s["Start Time (s)"],
                    "End Time (s)": s["End Time (s)"],
                    "Duration (s)": s["Duration (s)"],
                    transcript_col: txt,
                    "Intent": _tag(txt, _IR) if _tag else "",
                    "Sentiment": _tag(txt, _SR) if _tag else "",
                    "Outcome": _tag(txt, _OR) if _tag else "",
                    "Language": s["Language"],
                    "Audio File": s["Audio File"]
                }
                final_segments.append(row)

            columns = ["Segment #", "Speaker", "Start Time (s)", "End Time (s)", "Duration (s)", transcript_col, "Intent", "Sentiment", "Outcome", "Language", "Audio File"]

            # GPT-4o Call Insights on the corrected full transcript.
            try:
                from processor import get_call_intelligence as _gci
                _full = "\n".join(
                    f"{s['Speaker']}: {s['Transcript']}" for s in segments
                )
                _ci = _gci(_full) if _full.strip() else {}
            except Exception as _ce:
                print(f"[export] Call insights generation failed: {_ce}")
                _ci = {}

            def _join(v):
                if isinstance(v, (list, tuple)):
                    return ", ".join(str(x) for x in v) if v else "None"
                return str(v) if v not in (None, "") else "—"

            call_insights = [
                ("Audio File", original_filename),
                ("Language", ", ".join(sorted(list(languages))) or lang_str),
                ("Call Intent", _join(_ci.get("intent"))),
                ("Customer Mood", _join(_ci.get("mood"))),
                ("Churn Risk", _join(_ci.get("churn_risk"))),
                ("Onboarding Friction", _join(_ci.get("onboarding_friction"))),
                ("Operational Pain", _join(_ci.get("operational_pain"))),
                ("Financial Disputes", _join(_ci.get("financial_disputes"))),
                ("Service Leakage", _join(_ci.get("service_leakage"))),
                ("Summary", _join(_ci.get("summary"))),
            ]
            
            total_duration_secs = sum(s["Duration (s)"] for s in segments)
            unique_speakers = sorted(list(set(s["Speaker"] for s in segments)))
            summary_headers = ["Language", "Total Segments", "Total Duration (mins)", "Speakers", "Annotated By", "Export Date"]
            summary_values_list = [[
                ", ".join(sorted(list(languages))),
                len(segments),
                round(total_duration_secs / 60, 2),
                ", ".join(unique_speakers),
                ", ".join(sorted(list(annotators))) or "Vaidik AI" if internal_export else "Vaidik AI",
                datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ]]

        df_transcript = pd.DataFrame(final_segments)[columns] if final_segments else pd.DataFrame(columns=columns)

        # STEP 3: Create XLSX with premium formatting
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        stem = Path(original_filename).stem
        xlsx_filename = f"{stem}_annotated_{timestamp}.xlsx"
        xlsx_path = f"/tmp/vaidikai/{client_code}/{xlsx_filename}"
        os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

        wb = Workbook()
        ws1 = wb.active
        sheet_title = "Transcript - Annotated"
        if is_image_project:
            sheet_title = "Inventory Data"
            if is_housing:
                sheet_title = "Housing Collateral Data"
            elif is_business:
                sheet_title = "Business Audit Data"
        ws1.title = sheet_title
        
        HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        BODY_FONT = Font(name="Arial", size=10)
        ALTERNATE_FILL = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
        ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
        ALIGN_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)

        for col_num, header in enumerate(columns, 1):
            cell = ws1.cell(row=1, column=col_num, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = ALIGN_CENTER

        SPEAKER_COLORS = {
            "AI": Font(name="Arial", size=10, bold=True, color="000080"), # Dark Blue
            "Speaker 1": Font(name="Arial", size=10, bold=True, color="006400"), # Dark Green
            "Speaker 2": Font(name="Arial", size=10, bold=True, color="8B0000"), # Dark Red
        }

        for row_num, row_data in enumerate(df_transcript.values, 2):
            is_alternate = row_num % 2 != 0
            for col_num, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_num, column=col_num, value=value)
                
                # Apply speaker-specific font if this is the Speaker column (col 2)
                if columns[col_num-1] == "Speaker" and value in SPEAKER_COLORS:
                    cell.font = SPEAKER_COLORS[value]
                else:
                    cell.font = BODY_FONT
                
                if is_alternate:
                    cell.fill = ALTERNATE_FILL
                
                if columns[col_num-1] in ["Segment #", "Item #", "Start Time (s)", "End Time (s)", "Duration (s)", "Points Count"]:
                    cell.alignment = ALIGN_CENTER
                else:
                    cell.alignment = ALIGN_LEFT

        widths = {
            "Item #": 10, "Category": 20, "Geometry Type": 15, "Points Count": 12, "Annotator": 20, "Image File": 30,
            "Segment #": 10, "Speaker": 15, "Start Time (s)": 15, "End Time (s)": 15, "Duration (s)": 15, "Transcript": 60,
            "Intent": 20, "Sentiment": 14, "Outcome": 24, "Language": 12, "Audio File": 30,
            "AI Summary": 50, "Story Gist": 60
        }
        for col_num, header in enumerate(columns, 1):
            w = widths.get(header, 15)
            if isinstance(header, str) and header.startswith("Transcript"):
                w = 70
            ws1.column_dimensions[get_column_letter(col_num)].width = w

        # Call Insights sheet (audio only) — placed between Transcript and Summary.
        if call_insights:
            wsi = wb.create_sheet(title="Call Insights")
            t = wsi.cell(row=1, column=1, value="CALL INTELLIGENCE REPORT")
            t.fill = HEADER_FILL
            t.font = HEADER_FONT
            t.alignment = Alignment(horizontal="left", vertical="center")
            wsi.cell(row=1, column=2).fill = HEADER_FILL
            SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            LABEL_FONT = Font(name="Arial", size=10, bold=True)
            for idx, (k, v) in enumerate(call_insights, start=3):
                kc = wsi.cell(row=idx, column=1, value=k)
                kc.font = LABEL_FONT
                kc.fill = SECTION_FILL
                kc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                vc = wsi.cell(row=idx, column=2, value=v)
                vc.font = BODY_FONT
                vc.alignment = ALIGN_LEFT
                wsi.row_dimensions[idx].height = 70 if k == "Summary" else (40 if len(str(v)) > 60 else 18)
            wsi.column_dimensions["A"].width = 24
            wsi.column_dimensions["B"].width = 95

        # Segments sheet (clickstream only) — 3 segmentation axes side-by-side
        # so leadership can see who's hitting friction broken down by User Type,
        # Device-cohort (release impact), and Behavioural Archetype.
        if clickstream_segments_breakdown:
            wseg = wb.create_sheet(title="Segments")
            SEG_HEADER_FILL = PatternFill(start_color="2E5984", end_color="2E5984", fill_type="solid")
            SEG_SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
            SEG_LABEL_FONT = Font(name="Arial", size=10, bold=True)
            col_cursor = 1
            for axis_title, rows in clickstream_segments_breakdown.items():
                # Section header spanning 2 cols
                hc = wseg.cell(row=1, column=col_cursor, value=axis_title)
                hc.fill = SEG_HEADER_FILL
                hc.font = HEADER_FONT
                hc.alignment = ALIGN_CENTER
                wseg.cell(row=1, column=col_cursor + 1).fill = SEG_HEADER_FILL
                wseg.merge_cells(
                    start_row=1, start_column=col_cursor,
                    end_row=1, end_column=col_cursor + 1
                )
                # Sub-header
                sh1 = wseg.cell(row=2, column=col_cursor, value="Segment")
                sh2 = wseg.cell(row=2, column=col_cursor + 1, value="Count")
                for c in (sh1, sh2):
                    c.fill = SEG_SECTION_FILL
                    c.font = SEG_LABEL_FONT
                    c.alignment = ALIGN_CENTER
                # Rows
                for r_idx, (label, count) in enumerate(rows, start=3):
                    wseg.cell(row=r_idx, column=col_cursor, value=str(label or "—")).font = BODY_FONT
                    wseg.cell(row=r_idx, column=col_cursor + 1, value=count).font = BODY_FONT
                # Column widths
                wseg.column_dimensions[get_column_letter(col_cursor)].width = 32
                wseg.column_dimensions[get_column_letter(col_cursor + 1)].width = 10
                col_cursor += 3  # skip a gap column

        # By User sheet (clickstream only) — one row per unique hashed phone,
        # with the full session arc, archetype, recovery flag, and persistent
        # friction signal. This is the headline artefact for leadership: it
        # answers "which users are struggling" not just "which sessions failed".
        if clickstream_by_user_rows:
            wsu = wb.create_sheet(title="By User")
            user_cols = ["User ID (hashed)", "Total Sessions", "User Archetype",
                         "Recovery", "Persistent Friction", "User Type", "Device",
                         "Session Arc (chronological)"]
            for col_num, header in enumerate(user_cols, 1):
                c = wsu.cell(row=1, column=col_num, value=header)
                c.fill = HEADER_FILL
                c.font = HEADER_FONT
                c.alignment = ALIGN_CENTER
            for row_num, urow in enumerate(clickstream_by_user_rows, 2):
                is_alt = row_num % 2 != 0
                for col_num, key in enumerate(user_cols, 1):
                    c = wsu.cell(row=row_num, column=col_num, value=urow.get(key, ""))
                    c.font = BODY_FONT
                    c.alignment = ALIGN_LEFT
                    if is_alt:
                        c.fill = ALTERNATE_FILL
            user_widths = {"User ID (hashed)": 36, "Total Sessions": 14, "User Archetype": 22,
                           "Recovery": 10, "Persistent Friction": 30, "User Type": 12,
                           "Device": 28, "Session Arc (chronological)": 80}
            for col_num, header in enumerate(user_cols, 1):
                wsu.column_dimensions[get_column_letter(col_num)].width = user_widths.get(header, 18)

        ws2 = wb.create_sheet(title="Summary")
        for col_num, header in enumerate(summary_headers, 1):
            cell = ws2.cell(row=1, column=col_num, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = ALIGN_CENTER
        
        for row_idx, row_vals in enumerate(summary_values_list, 2):
            for col_num, value in enumerate(row_vals, 1):
                cell = ws2.cell(row=row_idx, column=col_num, value=value)
                cell.font = BODY_FONT
                if summary_headers[col_num-1] in ["Speakers", "Annotated By"]:
                    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                else:
                    cell.alignment = ALIGN_LEFT

        for col_num in range(1, len(summary_headers) + 1):
            ws2.column_dimensions[get_column_letter(col_num)].width = 25

        blob_service_client = BlobServiceClient.from_connection_string(connection_string)
        
        if is_image_project:
            # 1) Setup base dir
            base_dir = f"/tmp/vaidikai/{client_code}/{stem}_export_{timestamp}"
            audit_dir = os.path.join(base_dir, "Audit")
            ml_dir = os.path.join(base_dir, "ML_Training")
            os.makedirs(audit_dir, exist_ok=True)
            os.makedirs(ml_dir, exist_ok=True)
            
            # Save excel to Audit
            xlsx_name_base = "inventory"
            if is_housing:
                xlsx_name_base = "housing_verification"
            elif is_business:
                xlsx_name_base = "business_audit"
            wb.save(os.path.join(audit_dir, f"{stem}_{xlsx_name_base}_{timestamp}.xlsx"))
            
            # 2) Download original image from Azure (dynamically matching timestamp-prefixed blobs)
            intake_cc = blob_service_client.get_container_client("client-intake")
            blobs = list(intake_cc.list_blobs(name_starts_with=f"{client_code}/"))
            matching = [b.name for b in blobs if b.name.endswith(original_filename)]
            
            if not matching:
                raise ValueError(f"No blob found in client-intake matching {client_code}/*{original_filename}")
            src_blob_name = sorted(matching)[-1]
            
            intake_blob_client = blob_service_client.get_blob_client(container="client-intake", blob=src_blob_name)
            img_path = os.path.join(base_dir, f"raw_{original_filename}")
            with open(img_path, "wb") as f:
                f.write(intake_blob_client.download_blob().readall())
            
            img = cv2.imread(img_path)
            if img is None:
                raise ValueError(f"Failed to read image from {img_path}")
            h, w = img.shape[:2]
            
            # Apply dynamic face redaction blurring for borrower or representative faces
            from redactor import redact_faces_in_image
            img = redact_faces_in_image(img, segments)
            
            # ── DUPLICATE COLLATERAL DETECTION ──────────────────────
            # Generate signatures for each image label item
            new_signatures = []
            for item in segments:
                pts = item.get("Points", [])
                if item.get("RType") in ("polygonlabels", "rectanglelabels") and pts:
                    sig = generate_signature(
                        image=img,
                        polygon_points=pts,
                        category=item["Category"],
                        client_code=client_code,
                        task_id=matched_tasks[0].get("id", 0) if matched_tasks else 0,
                        item_index=item["Item #"],
                        image_file=original_filename,
                        img_width=w,
                        img_height=h,
                    )
                    if sig:
                        new_signatures.append(sig)
            
            # Check for duplicates against historical database
            if new_signatures and not force_delivery:
                matches = find_duplicates(new_signatures)
                if matches:
                    # Clean up temp files
                    os.remove(img_path)
                    shutil.rmtree(base_dir, ignore_errors=True)
                    
                    return {
                        "status": "duplicate_warning",
                        "matches": matches,
                        "total_matches": len(matches),
                        "message": f"⚠️ COLLATERAL MATCH DETECTED: {len(matches)} item(s) match historical records. Review required before delivery.",
                    }
            
            print(f"[Collateral Detector] No duplicates found. Proceeding with delivery.")
            
            # 3) Generate COCO JSON & draw polygons
            coco = {
                "images": [{"id": 1, "width": w, "height": h, "file_name": original_filename}],
                "annotations": [],
                "categories": []
            }
            cat_map = {}
            
            supercat = "jewelry"
            if is_housing:
                supercat = "housing"
            elif is_business:
                supercat = "msme_business"
                
            for item in segments:
                cat = item["Category"]
                if cat not in cat_map:
                    cat_map[cat] = len(cat_map) + 1
                    coco["categories"].append({"id": cat_map[cat], "name": cat, "supercategory": supercat})
                
                pts = item.get("Points", [])
                if item.get("RType") in ("polygonlabels", "rectanglelabels") and pts:
                    # Draw on image
                    cv_pts = np.array([[int(p[0] * w / 100.0), int(p[1] * h / 100.0)] for p in pts], np.int32)
                    cv_pts = cv_pts.reshape((-1, 1, 2))
                    cv2.polylines(img, [cv_pts], True, (0, 255, 0), 3)
                    
                    # Add label
                    cx, cy = cv_pts[0][0]
                    cv2.putText(img, f"#{item['Item #']} {cat}", (cx, cy-10), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 2)
                    
                    # Add to COCO
                    coco_pts = [p for pt in pts for p in [pt[0] * w / 100.0, pt[1] * h / 100.0]]
                    
                    # Calculate real bounding box dynamically
                    xs = [pt[0] * w / 100.0 for pt in pts]
                    ys = [pt[1] * h / 100.0 for pt in pts]
                    xmin = min(xs)
                    ymin = min(ys)
                    xmax = max(xs)
                    ymax = max(ys)
                    bbox_w = xmax - xmin
                    bbox_h = ymax - ymin
                    bbox = [round(xmin, 2), round(ymin, 2), round(bbox_w, 2), round(bbox_h, 2)]
                    
                    # Shoelace formula for polygon area calculation
                    pixel_pts = [(pt[0] * w / 100.0, pt[1] * h / 100.0) for pt in pts]
                    n = len(pixel_pts)
                    polygon_area = 0.0
                    for idx in range(n):
                        next_idx = (idx + 1) % n
                        polygon_area += pixel_pts[idx][0] * pixel_pts[next_idx][1]
                        polygon_area -= pixel_pts[next_idx][0] * pixel_pts[idx][1]
                    polygon_area = abs(polygon_area) / 2.0
                    
                    coco["annotations"].append({
                        "id": item["Item #"],
                        "image_id": 1,
                        "category_id": cat_map[cat],
                        "segmentation": [coco_pts],
                        "area": round(polygon_area, 2),
                        "bbox": bbox,
                        "iscrowd": 0
                    })
            
            # Save image
            cv2.imwrite(os.path.join(audit_dir, f"{stem}_annotated.jpg"), img)
            os.remove(img_path)
            
            # Save JSON
            with open(os.path.join(ml_dir, f"{stem}_coco.json"), "w") as f:
                json.dump(coco, f, indent=2)
            
            # Zip it all up
            zip_filename = f"{stem}_export_{timestamp}.zip"
            zip_path = f"/tmp/vaidikai/{client_code}/{stem}_export_{timestamp}"
            shutil.make_archive(zip_path, 'zip', base_dir)
            
            # Upload ZIP
            delivery_blob_name = f"{client_code}/{zip_filename}"
            blob_client = blob_service_client.get_blob_client(container="client-delivery", blob=delivery_blob_name)
            with open(zip_path + ".zip", "rb") as data:
                blob_client.upload_blob(data, overwrite=True)
                
            shutil.rmtree(base_dir, ignore_errors=True)
            try:
                os.remove(zip_path + ".zip")
            except:
                pass
                
            # Store signatures after successful delivery
            if new_signatures:
                store_signatures(new_signatures)
            
            return {
                "status": "success",
                "rows_exported": len(segments),
                "xlsx_filename": zip_filename,
                "delivery_path": f"client-delivery/{client_code}/",
                "message": f"Successfully exported Gold Standard Bundle to {zip_filename}"
            }
            
        else:
            wb.save(xlsx_path)
            print(f"XLSX saved to {xlsx_path}")

            # STEP 4: Upload to Azure
            delivery_blob_name = f"{client_code}/{xlsx_filename}"
            blob_client = blob_service_client.get_blob_client(container="client-delivery", blob=delivery_blob_name)
            
            with open(xlsx_path, "rb") as data:
                blob_client.upload_blob(data, overwrite=True)
            print(f"Uploaded to client-delivery/{delivery_blob_name}")
            os.remove(xlsx_path)
            try:
                parent_dir = os.path.dirname(xlsx_path)
                if not os.listdir(parent_dir):
                    os.rmdir(parent_dir)
            except Exception:
                pass

            return {
                "status": "success",
                "rows_exported": len(segments),
                "xlsx_filename": xlsx_filename,
                "delivery_path": f"client-delivery/{client_code}/",
                "message": f"Successfully exported {len(segments)} segments to XLSX"
            }

    except Exception as e:
        error_msg = f"Error in export and delivery: {str(e)}"
        print(error_msg)
        return {
            "status": "error",
            "error": error_msg,
            "client_code": client_code,
            "original_filename": original_filename
        }

_TASK_CACHE = {}

def check_annotation_status(
    client_code: str,
    project_id: str,
    original_filename: str = None,
) -> Dict[str, Any]:
    """
    Check the annotation status for a client (or specific file) in Label Studio across all project IDs with 15s TTL caching.
    """
    import time
    global _TASK_CACHE
    try:
        ls_url = os.getenv("LABEL_STUDIO_URL", "").rstrip("/")
        headers = _get_ls_headers()

        project_ids_to_check = [str(project_id)]
        for pid in ["1", "2", "3", "4", "5", "6", "7"]:
            if pid not in project_ids_to_check:
                project_ids_to_check.append(pid)

        filtered_tasks = []
        actual_pid = project_id

        for pid in project_ids_to_check:
            try:
                now = time.time()
                cache_key = f"{ls_url}_{pid}"
                if cache_key in _TASK_CACHE and now - _TASK_CACHE[cache_key]["time"] < 15:
                    all_tasks = _TASK_CACHE[cache_key]["tasks"]
                else:
                    r = requests.get(
                        f"{ls_url}/api/tasks",
                        params={"project": pid, "page_size": 1000},
                        headers=headers,
                        timeout=10
                    )
                    if r.status_code == 200:
                        resp = r.json()
                        all_tasks = resp if isinstance(resp, list) else resp.get("tasks", [])
                        _TASK_CACHE[cache_key] = {"tasks": all_tasks, "time": now}
                    else:
                        all_tasks = []

                tasks_for_file = [
                    t for t in all_tasks
                    if t.get("data", {}).get("client_code") == client_code
                    and (original_filename is None or t.get("data", {}).get("filename") == original_filename)
                ]
                if tasks_for_file:
                    filtered_tasks = tasks_for_file
                    actual_pid = pid
                    break
            except Exception:
                pass

        total = len(filtered_tasks)
        completed = sum(1 for t in filtered_tasks if t.get("total_annotations", 0) > 0)

        return {
            "client_code": client_code,
            "filename": original_filename,
            "total_segments": total,
            "completed_annotations": completed,
            "pending_annotations": total - completed,
            "completion_percentage": round((completed / total * 100), 1) if total > 0 else 0,
            "project_id": actual_pid,
        }

    except Exception as e:
        return {
            "status": "error",
            "error": str(e),
            "client_code": client_code,
        }
